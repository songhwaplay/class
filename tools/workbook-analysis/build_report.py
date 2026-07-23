from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, range_boundaries
from openpyxl.formula.tokenizer import Tokenizer


ROOT = Path(__file__).resolve().parent
ANALYSIS = ROOT / "_analysis" / "연산" / "analysis.json"
OUTPUT = ROOT / "연산_XLSM_분석보고서.md"
A = json.loads(ANALYSIS.read_text(encoding="utf-8"))


def esc(value) -> str:
    text = "" if value is None else str(value)
    return text.replace("|", "\\|").replace("\r", " ").replace("\n", "<br>")


def code(value) -> str:
    return f"`{esc(value)}`"


def md_table(headers, rows) -> list[str]:
    out = ["| " + " | ".join(headers) + " |", "| " + " | ".join("---" for _ in headers) + " |"]
    out.extend("| " + " | ".join(esc(x) for x in row) + " |" for row in rows)
    return out


def compress_cells(refs: list[str]) -> str:
    coords = set()
    others = []
    for ref in refs:
        try:
            col, row = coordinate_from_string(ref.replace("$", ""))
            coords.add((column_index_from_string(col), row))
        except Exception:
            others.append(ref)
    groups = []
    by_col = defaultdict(list)
    for col, row in sorted(coords):
        by_col[col].append(row)
    for col, rows in by_col.items():
        start = prev = rows[0]
        for row in rows[1:] + [None]:
            if row is not None and row == prev + 1:
                prev = row
                continue
            c = get_column_letter(col)
            groups.append(f"{c}{start}" if start == prev else f"{c}{start}:{c}{prev}")
            if row is not None:
                start = prev = row
    groups.extend(others)
    return ", ".join(groups)


def anchor_to_range(item: dict) -> str:
    raw = item.get("Anchor")
    if raw:
        nums = [int(x) for x in re.findall(r"-?\d+", raw)]
        if len(nums) >= 8:
            return f"{get_column_letter(nums[0] + 1)}{nums[2] + 1}:{get_column_letter(nums[4] + 1)}{nums[6] + 1}"
    pos = item.get("position") or {}
    if "from" in pos:
        f = pos["from"]
        t = pos.get("to", f)
        return f"{get_column_letter(int(f.get('col', 0)) + 1)}{int(f.get('row', 0)) + 1}:{get_column_letter(int(t.get('col', f.get('col', 0))) + 1)}{int(t.get('row', f.get('row', 0))) + 1}"
    return ""


def formula_functions(formula: str) -> list[str]:
    return [x.upper() for x in re.findall(r"(?<![A-Za-z0-9_.])([A-Za-z_][A-Za-z0-9_.]*)\s*\(", formula or "")]


SHEET_NAMES = [s["name"] for s in A["sheets"]]


def formula_sources(formula: str) -> list[str]:
    found = []
    f = formula or ""
    for name in SHEET_NAMES:
        quoted = "'" + name.replace("'", "''") + "'!"
        bare = name + "!"
        if quoted in f or bare in f:
            found.append(name)
    return found


def local_formula_ref_cells(formula: str) -> set[str]:
    refs = set()
    try:
        items = Tokenizer("=" + (formula or "")).items
    except Exception:
        return refs
    for item in items:
        if item.type != "OPERAND" or item.subtype != "RANGE" or "!" in item.value:
            continue
        value = item.value.replace("$", "")
        if not re.fullmatch(r"[A-Z]{1,3}\d+(?::[A-Z]{1,3}\d+)?", value, re.I):
            continue
        try:
            min_col, min_row, max_col, max_row = range_boundaries(value)
        except Exception:
            continue
        if (max_col - min_col + 1) * (max_row - min_row + 1) > 5000:
            continue
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                refs.add(f"{get_column_letter(col)}{row}")
    return refs


PROC_MAP = defaultdict(list)
for module in A["vba"]["modules"]:
    for proc in module["procedures"]:
        PROC_MAP[proc["name"].casefold()].append((module, proc))


def macro_name(raw: str) -> str:
    return raw.split("!", 1)[-1].strip("'\"")


def range_args(body: str) -> list[str]:
    vals = []
    for m in re.finditer(r"(?is)\bRange\s*\((.*?)\)", body):
        strings = re.findall(r'"([^"]*)"', m.group(1))
        if strings:
            value = "".join(strings)
            if value not in vals:
                vals.append(value)
    return vals


def proc_effect(body: str) -> str:
    effects = []
    checks = [
        (r"\bCalculate\b", "전체 재계산"),
        (r"\bRandomize\b|\bRnd\b", "난수 생성"),
        (r"\.Font\b|\.Color\b|ThemeColor", "글자색으로 문제/답 표시 전환"),
        (r"\.Borders\b|FormatConditions", "테두리/조건부서식 변경"),
        (r"Cells\.Clear|\.Cells\.Clear", "데이터 시트 초기화"),
        (r"\.Value\s*=|Cells\s*\([^\)]*\)\s*=", "셀 값 기록"),
        (r"Sheets\.Add", "시트 생성"),
        (r"\.Unprotect|\.Protect", "시트 보호 전환"),
        (r"xlDialogPrint", "인쇄 대화상자"),
        (r"\.Show\b", "폼 표시"),
        (r"\.Select\b", "시트/범위 선택"),
        (r"Application\.Calculation", "계산 모드 변경"),
    ]
    for pattern, label in checks:
        if re.search(pattern, body, re.I):
            effects.append(label)
    return ", ".join(effects) or "보조 계산"


def button_assignments():
    rows = []
    for sheet in A["sheets"]:
        for kind, items in [("폼 버튼", sheet["vml_shapes"]), ("도형/그림", sheet["drawings"])]:
            for item in items:
                raw = item.get("FmlaMacro") or item.get("Macro") or item.get("macro")
                if not raw:
                    continue
                macro = macro_name(raw)
                resolved = PROC_MAP.get(macro.casefold(), [])
                rows.append(
                    {
                        "sheet": sheet["name"],
                        "code": sheet["codeName"],
                        "kind": kind,
                        "object": item.get("id") or item.get("name") or "",
                        "anchor": anchor_to_range(item),
                        "macro": macro,
                        "resolved": resolved,
                    }
                )
    return rows


BUTTONS = button_assignments()


def build() -> str:
    lines = []
    lines += [
        "# 연산.xlsm 정적 분석 보고서",
        "",
        "> 분석 방식: 원본을 저장하거나 수정하지 않고 XLSM의 Open XML, VML/도형 관계, 프린터 바이너리, VBA OLE 스트림을 읽기 전용으로 분석했다. 매크로는 실행하지 않았다.",
        "",
        "## 1. 결론",
        "",
        "이 통합문서는 일반적인 셀 입력형 계산기가 아니라, 초등 수학 문제를 난수로 생성하고 정답을 숨기거나 표시하며 인쇄하는 워크시트형 문제 생성기다. 수학 로직, 문제 재생성, 정답 표시, 시트 선택, 인쇄는 웹으로 구현할 수 있다. 다만 Excel의 전역 재계산, 셀 글자색을 이용한 정답 은폐, UserForm, 보호 시트, 인쇄 설정, 시계/도형·필기 영역은 웹 상태 모델·SVG/Canvas·인쇄 CSS로 재설계해야 한다.",
        "",
        "정적 분석에서 이미 확인되는 끊어진 연결이 있다: 매크로가 존재하지 않는 버튼 배정 11건, 존재하지 않는 시트를 가리키는 UserForm 이벤트 4건, 존재하지 않는 `결과입력` 폼 호출 1건, `#NAME?`인 숨김 정의 이름 1개, 숨김 `연습일지`의 캐시된 `#VALUE!` 수식 39개다. 동일 동작을 웹에 옮기기 전에 이 항목들의 의도를 확정해야 한다.",
        "",
        "## 2. 파일 및 구조 요약",
        "",
    ]
    lines += md_table(
        ["항목", "값"],
        [
            ["원본", A["source"]],
            ["SHA-256", A["sha256"]],
            ["크기", f"{A['size']:,} bytes"],
            ["워크시트", f"81개: 보임 64, 숨김 17, veryHidden 0"],
            ["통합문서 구조 보호", "lockStructure=1, SHA-512 해시"],
            ["내용 셀", f"{sum(len(s['cells']) for s in A['sheets']):,}"],
            ["수식 셀", f"{sum(len(s['formulas']) for s in A['sheets']):,}"],
            ["정의 이름", f"{len(A['defined_names'])}개"],
            ["Excel 표", "0개"],
            ["데이터 유효성 검사", "0개"],
            ["조건부서식", f"{sum(len(s.get('conditional_formats', [])) for s in A['sheets'])}개 범위 / {sum(sum(len(cf['rules']) for cf in s.get('conditional_formats', [])) for s in A['sheets'])}개 규칙"],
            ["폼 컨트롤", "208개"],
            ["매크로 연결 도형/그림", f"{sum(1 for b in BUTTONS if b['kind']=='도형/그림')}개"],
            ["프린터 설정", f"{len(A['printer_parts'])}개"],
            ["VBA", "표준 모듈 49, 문서 모듈 82, UserForm 1, 클래스 모듈 0, 프로시저 166"],
        ],
    )

    lines += ["", "## 3. 시트·수식·컨트롤 전수 목록", "", "수식 예시는 각 시트의 첫 수식들이다. `난수 셀`은 `RAND` 또는 `RANDBETWEEN`을 포함하는 수식 주소이며, 이 값이 해당 시트의 문제와 정답을 함께 구동한다.", ""]
    sheet_rows = []
    for s in A["sheets"]:
        funcs = Counter()
        sources = Counter()
        random_cells = []
        for c in s["formulas"]:
            funcs.update(formula_functions(c["formula"] or ""))
            sources.update(formula_sources(c["formula"] or ""))
            if re.search(r"\bRAND(?:BETWEEN)?\s*\(", c["formula"] or "", re.I):
                random_cells.append(c["ref"])
        macros = []
        for b in BUTTONS:
            if b["sheet"] == s["name"] and b["macro"] not in macros:
                macros.append(b["macro"])
        examples = [f"{c['ref']}={c['formula']}" for c in s["formulas"][:3]]
        sheet_rows.append(
            [
                s["index"] + 1,
                s["name"],
                s["state"],
                s["codeName"],
                s["dimension"],
                len(s["formulas"]),
                ", ".join(f"{k}:{v}" for k, v in funcs.most_common()),
                ", ".join(f"{k}({v})" for k, v in sources.items()),
                compress_cells(random_cells),
                ", ".join(macros),
                "<br>".join(examples),
            ]
        )
    lines += md_table(["#", "시트", "상태", "코드명", "사용 범위", "수식", "함수", "시트 선행값", "난수 셀", "사용자 동작", "수식 근거 예"], sheet_rows)

    hidden_dependents = defaultdict(Counter)
    for s in A["sheets"]:
        for c in s["formulas"]:
            for source in formula_sources(c["formula"] or ""):
                hidden_dependents[source][s["name"]] += 1
    lines += ["", "## 4. 숨김 시트와 표시 시트의 의존관계", ""]
    hidden_rows = []
    for s in A["sheets"]:
        if s["state"] != "hidden":
            continue
        deps = hidden_dependents.get(s["name"], {})
        hidden_rows.append(
            [s["name"], s["codeName"], s["dimension"], len(s["formulas"]), ", ".join(f"{k}({v}셀)" for k, v in deps.items()) or "정적 수식 종속 없음", "VBA 기록 대상" if s["name"] in {"분수대소데이터", "6자연수원본"} else "수식/정적 데이터 원본"]
        )
    lines += md_table(["숨김 시트", "코드명", "범위", "수식", "표시 시트 종속", "역할"], hidden_rows)
    lines += [
        "",
        "주요 흐름은 다음과 같다.",
        "",
        "```text",
        "[새 문제 버튼] -> Module4.문제바꾸기 -> Excel 전체 Calculate",
        "    -> RAND/RANDBETWEEN 재평가",
        "    -> 숨김 원본의 RANK/INDEX 순서 재편성",
        "    -> 표시 시트 문제 셀과 정답 셀 동시 갱신",
        "    -> *_문제 프로시저: 정답 범위 글자색/테두리 숨김",
        "    -> *_답 프로시저: 같은 범위 글자색/테두리 표시",
        "```",
        "",
        "`Calculate`가 시트나 범위로 한정되지 않았기 때문에 한 시트의 새 문제 버튼을 눌러도 통합문서 전체의 휘발성 수식이 재계산된다. 웹에서는 이 전역 동작을 그대로 복제할지, 현재 문제만 재생성할지 결정해야 한다.",
        "",
        "### 특수 VBA 데이터 생성 흐름",
        "",
    ]
    lines += md_table(
        ["사용자 동작", "VBA 근거", "중간 출력", "표시 출력"],
        [
            ["5분수③ 데이터 새로 생성", "Module10.fifth분수대소데이터, 2-135행; GetVbaGcd 137-142행; RandBetween 144-146행", "분수대소데이터!A1:H11을 지우고 10개 비교 문제 기록", "5분수③!C5:E32 등 50개 직접 참조 수식"],
            ["6소수① 새로고침", "Module1.sixth소수A_새로고침, 38-130행", "6자연수원본!A1:D19을 지우고 18개 나눗셈·정답 기록", "6소수①!C5:I35의 54개 직접 참조 수식"],
            ["문제 선택", "Module35.문제지선택, 3-5행", "UserForm 문제선택 표시", "각 Click 이벤트가 ThisWorkbook.Sheets(이름).Select 후 폼 닫기"],
            ["인쇄", "Module20.인쇄, 2-6행", "Application.Dialogs(xlDialogPrint).Show", "Excel 인쇄 대화상자"],
        ],
    )

    lines += ["", "## 5. 셀 입력과 출력 의존성", ""]
    unlocked_rows = []
    for s in A["sheets"]:
        unlocked = [c for c in s.get("all_cells", []) if not c.get("locked", True)]
        if not unlocked:
            continue
        blanks = [c["ref"] for c in unlocked if c.get("value") is None and c.get("formula") is None]
        values = [f"{c['ref']}={c.get('value')!r}" for c in unlocked if c.get("value") is not None and c.get("formula") is None]
        refs_by_formulas = []
        for c in s["formulas"]:
            local_refs = local_formula_ref_cells(c["formula"] or "")
            for u in unlocked:
                if u["ref"] in local_refs:
                    refs_by_formulas.append(f"{c['ref']}<-{u['ref']}")
        unlocked_rows.append([s["name"], compress_cells(blanks), "; ".join(values), "; ".join(refs_by_formulas) or "해당 시트 수식에서 직접 참조되지 않음"])
    lines += md_table(["시트", "잠금 해제 빈 셀", "잠금 해제 값 셀", "수식 종속"], unlocked_rows)
    lines += [
        "",
        "따라서 대부분의 사용자는 셀 값을 넣지 않고 버튼을 입력으로 사용한다. `2시계①`의 18개 잠금 해제 값 셀과 `3분수②!H4:I26, K4:K26`의 잠금 해제 작업 영역은 로컬 수식에 직접 연결되지 않는다. 웹에서는 이 영역을 답안 입력 또는 필기 캔버스로 만들지, 단순 표시로 둘지 별도 결정이 필요하다.",
        "",
        "## 6. 이름 정의, 표, 데이터 유효성, 조건부서식",
        "",
    ]
    lines += md_table(["종류", "결과", "근거"], [["정의 이름", "_xleta.T 1개", "숨김=1, xlm=1, 정의식=#NAME?"], ["Excel 표", "없음", "xl/tables 파트와 시트 table 관계가 모두 없음"], ["데이터 유효성", "없음", "81개 시트의 dataValidation 요소 0개"], ["조건부서식", "8개 범위, 10개 규칙", "일부 문제/답 프로시저가 추가·삭제·테두리 변경"]])
    if any(s.get("conditional_formats") for s in A["sheets"]):
        cf_rows = []
        for s in A["sheets"]:
            for cf in s.get("conditional_formats", []):
                for rule in cf["rules"]:
                    cf_rows.append([s["name"], cf["sqref"], rule.get("type"), ", ".join(rule.get("formulas", [])), rule.get("priority")])
        lines += [""] + md_table(["시트", "범위", "형식", "수식", "우선순위"], cf_rows)

    lines += ["", "## 7. VBA 전수 조사", "", "### 모듈 구성", ""]
    lines += md_table(
        ["구분", "수", "확인 결과"],
        [
            ["표준 모듈", 49, "문제/답 표시, 전체 재계산, 특수 난수 데이터 생성, 인쇄, 폼 표시"],
            ["시트 모듈", 81, "Sheet27만 활성 이벤트 코드가 있고 나머지는 선언부만 존재"],
            ["ThisWorkbook", 1, "Workbook_BeforeSave가 주석 처리되어 자동 이벤트는 없음"],
            ["클래스 모듈", 0, "없음"],
            ["UserForm", 1, "문제선택; Caption 문자열은 2026-계상초등학교"],
        ],
    )
    lines += ["", "### 표준 모듈 및 활성 문서 프로시저", ""]
    proc_rows = []
    for m in A["vba"]["modules"]:
        if m["type"] not in {"standard", "document"}:
            continue
        for p in m["procedures"]:
            proc_rows.append([m["type"], m["name"], p["name"], f"{p['start_line']}-{p['end_line']}", proc_effect(p["body"]), "; ".join(range_args(p["body"])[:8])])
    lines += md_table(["모듈 종류", "모듈", "프로시저", "코드 행", "실제 효과", "주요 Range 근거"], proc_rows)

    lines += ["", "### UserForm 문제선택 이벤트", ""]
    form_rows = []
    for m in A["vba"]["modules"]:
        if m["type"] != "userform":
            continue
        for p in m["procedures"]:
            targets = re.findall(r"(?i)(?:ThisWorkbook\.)?Sheets\s*\(\s*\"([^\"]+)\"\s*\)", p["body"])
            form_rows.append([p["name"], f"{p['start_line']}-{p['end_line']}", ", ".join(targets), "정상" if all(t in SHEET_NAMES for t in targets) else "대상 시트 없음"])
    lines += md_table(["이벤트", "코드 행", "선택 시트", "정적 판정"], form_rows)

    lines += ["", "## 8. 사용자 버튼과 실제 호출 프로시저 전수 연결", "", "버튼 자체의 VML Caption은 비어 있어 표시문구는 패키지에서 얻을 수 없었다. 아래 `매크로`가 실제 동작 근거이며, 위치는 VML/도형 앵커를 셀 주소로 환산한 근사 범위다.", ""]
    button_rows = []
    for b in BUTTONS:
        target = ", ".join(f"{m['name']}.{p['name']}({p['start_line']}-{p['end_line']})" for m, p in b["resolved"])
        button_rows.append([b["sheet"], b["kind"], b["object"], b["anchor"], b["macro"], target or "없음", "정상" if target else "끊김"])
    lines += md_table(["시트", "객체", "ID/이름", "위치", "배정 매크로", "실제 프로시저", "판정"], button_rows)

    missing = [b for b in BUTTONS if not b["resolved"]]
    lines += ["", "## 9. 확인된 결함 또는 끊어진 연결", ""]
    defect_rows = []
    for b in missing:
        defect_rows.append(["버튼 매크로 없음", b["sheet"], f"{b['kind']} {b['anchor']}", b["macro"], "프로젝트 내 동명 프로시저가 없음"])
    defect_rows += [
        ["UserForm 대상 시트 없음", "문제선택", "C3시_Click 63-66", "3시간", "실제 시트는 3시간①/②이며 어느 쪽이 의도인지 미확정"],
        ["UserForm 대상 시트 없음", "문제선택", "C6소_Click 108-111", "6소수", "실제 시트는 6소수①/②/③이며 의도 미확정"],
        ["UserForm 대상 시트 없음", "문제선택", "CommandButton19_Click 183-186", "5인수분해", "동명 시트 없음; 자연수분해와의 관계는 추정 금지"],
        ["UserForm 대상 시트 없음", "문제선택", "CommandButton6_Click 228-231", "1수세기②", "동명 시트 없음"],
        ["UserForm 없음", "연습일지/Sheet27", "입력_Click 11-13", "결과입력.Show", "VBA 프로젝트에는 문제선택 UserForm만 존재"],
        ["정의 이름 오류", "통합문서", "_xleta.T", "#NAME?", "숨김 XLM 이름의 정의식이 오류"],
        ["캐시된 수식 오류", "연습일지", "G5:G43", "REPT(\"―\",E행-2)", "E5:E43이 비어 음수 반복수가 되어 #VALUE! 캐시"],
    ]
    lines += md_table(["종류", "위치", "근거", "대상", "판정"], defect_rows)

    lines += ["", "## 10. 외부 환경 의존성", ""]
    lines += md_table(
        ["환경", "확인 결과", "근거"],
        [
            ["외부 통합문서/연결", "없음", "externalLinks, connections, queryTables 파트 0; 외부 파일 참조 수식 없음"],
            ["운영 URL/하이퍼링크", "없음", "시트 hyperlink 0; VBA URL 문자열 0"],
            ["SharePoint 메타데이터", "잔존하지만 실행 의존성은 확인되지 않음", "customXml에 문서 콘텐츠 형식·DocumentLibraryForm·스키마 namespace가 있으나 외부 relationship은 없음"],
            ["파일·프로그램 실행", "없음", "Shell/CreateObject/GetObject/Workbooks.Open/Declare/FileCopy/Kill 호출 없음"],
            ["Office/VBA", "필수", "Excel8.0, EXCEL.EXE, MSO.DLL, VBE7.DLL, FM20.DLL, stdole2.tlb 참조"],
            ["UserForm", "Microsoft Forms 2.0", "문제선택/CompObj에 Microsoft Forms 2.0 Form"],
            ["프린터", "Microsoft Print to PDF로 저장됨", "69개 printerSettings 바이너리 모두 동일 장치명 포함"],
            ["인쇄 실행", "현재 Excel 환경의 인쇄 대화상자 사용", "Module20.인쇄의 xlDialogPrint"],
            ["통합문서/시트 보호", "구조 보호 및 대부분 시트 보호", "workbook lockStructure=1; 64개 보호 시트; Module8은 하드코딩된 비밀번호로 3나눗셈②를 해제/재보호"],
            ["기관/연도 문자열", "폼 Caption에 고정", "문제선택 VBFrame: 2026-계상초등학교"],
        ],
    )

    lines += [
        "",
        "## 11. 웹에서 동일 구현 가능한 기능",
        "",
        "- 사칙연산, 분수, 소수, 약수·배수, 단위 변환 등 문제 생성 수식과 VBA 난수 알고리즘",
        "- 새 문제, 문제 보기, 정답 보기, 문제 선택, 인쇄/PDF",
        "- 숨김 원본 시트의 데이터 선택·섞기·랭킹 로직",
        "- 문제별 정답 범위 표시와 단계별 풀이 영역",
        "- 보호 시트가 담당하던 읽기 전용/편집 가능 상태",
        "",
        "권장 구현 단위는 `활동 정의(JSON) + 순수 문제 생성 함수 + 정답 상태 + 렌더러`다. 난수는 seed를 저장해 문제를 재현할 수 있게 하고, 각 문제 생성 함수에 Excel 수식과 같은 제약을 테스트로 고정하는 편이 안전하다.",
        "",
        "## 12. 웹에서 재설계가 필요한 기능",
        "",
        "- 전역 `Calculate`: 현재 문제만 갱신하는 UX로 바꿀지 전체 문제를 동시에 갱신할지 결정",
        "- 글자색으로 정답 숨김: DOM 상태와 접근성 속성으로 전환",
        "- `Selection`/`ActiveSheet` 의존 VBA: 명시적인 활동 ID와 출력 영역으로 전환",
        "- UserForm: 라우팅 메뉴로 전환하고 존재하지 않는 4개 대상의 의도 확정",
        "- 시계·도형·WMF/EMF·Ink: SVG/Canvas로 재작성하고 실제 시각 검수",
        "- `3분수②!H4:I26, K4:K26`의 잠금 해제 작업 영역: 웹 필기 캔버스 또는 답안 입력으로 재정의",
        "- Microsoft Print to PDF: 브라우저 인쇄/PDF용 CSS 페이지 규격으로 재설계",
        "- 보호 비밀번호: 웹 보안 기능으로 취급하지 말고 권한/상태 제어로 대체",
        "- 7191개 수식을 브라우저에서 그대로 평가하기보다 문제 유형별 순수 함수로 이관하고 회귀 테스트 작성",
        "",
        "## 13. 실행해 봐야만 확정 가능한 항목",
        "",
        "- 겹쳐 있는 폼 버튼과 그림 중 실제로 위에 놓여 클릭되는 객체; 특히 `first덧셈b_답`이 배정된 6개 오래된 컨트롤의 사용자 영향",
        "- VML Caption이 비어 있는 버튼의 실제 표시문구와 이미지 레이어 조합",
        "- 현재 Excel 버전·매크로 보안 설정에서 VBA 프로젝트가 경고 없이 로드·실행되는지",
        "- 문제선택 UserForm의 실제 탭 순서, 크기, 잘림, 고해상도 표시",
        "- 69개 인쇄 레이아웃이 실제 용지에서 잘리지 않는지와 기본 프린터 변경 시 결과",
        "- RAND/RANDBETWEEN을 다시 계산했을 때 각 문제 제약이 항상 충족되는지; 특히 Module1의 제한 없는 Do While과 Module10의 5000회 탈출 경로",
        "- 보호 시트에서 문제/답 매크로의 서식 변경이 모든 Excel 버전에서 허용되는지",
        "- 시계, WMF/EMF, Ink 도형이 브라우저 변환 시 원본과 시각적으로 일치하는지",
        "",
        "이 목록은 추측 항목이 아니라 실행 또는 시각 상호작용 검사가 필요한 경계다. 원본 매크로는 이번 분석에서 실행하지 않았다.",
    ]
    return "\n".join(lines) + "\n"


if __name__ == "__main__":
    OUTPUT.write_text(build(), encoding="utf-8")
    print(OUTPUT)
